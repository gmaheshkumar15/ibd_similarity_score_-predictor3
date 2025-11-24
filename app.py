import streamlit as st
import pandas as pd
import joblib
from tensorflow.keras.models import load_model
from merge import merge_features
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO

# -----------------------------
# Load models safely
# -----------------------------
try:
    log_model = joblib.load("logistic_model_only.pkl")
    scaler = joblib.load("scaler_final1.pkl")
except Exception as e:
    st.error(f"Error loading models: {e}")

# -----------------------------
# Page config
# -----------------------------
st.set_page_config(page_title="IBD Risk Prediction", layout="wide")

# -----------------------------
# Custom CSS
# -----------------------------
st.markdown("""
    <style>
    .stApp { background-color: #ADD8E6; }
    .stSelectbox>div>div>div>select { text-align: center; } 
    .stSelectbox label {
        font-weight: bold !important;
        font-size: 22px !important;
        color: #000000 !important;
        text-align: center;
        width: 100%; 
        display: block; 
        margin-bottom: 5px; 
    }
    .stSelectbox select {
        border: 2px solid black; 
        border-radius: 5px; 
        padding: 5px 10px; 
    }
    .logo-left, .logo-right { width: 120px; display:block; margin:auto; }
    .institute-name { text-align:center; font-weight:bold; font-size:16px; margin-top:5px; }
    .large-score {
        font-size: 70px !important;
        font-weight: bold;
        color: #8B0000;
        text-align: center;
        margin-top: 20px;
    }
    .intro-paragraph {
        margin-bottom: 0px; 
        padding-bottom: 0px;
    }
    </style>
""", unsafe_allow_html=True)

# -----------------------------
# Logos and Title (Row 1)
# -----------------------------
col_logo_left, col_title, col_logo_right = st.columns([1, 5, 1])

with col_logo_left:
    st.markdown('<img src="https://brandlogovector.com/wp-content/uploads/2022/04/IIT-Delhi-Icon-Logo.png" class="logo-left">', unsafe_allow_html=True)
    st.markdown('<div class="institute-name">Indian Institute of Technology Delhi</div>', unsafe_allow_html=True)

with col_title:
    st.markdown(
        "<h1 style='text-align:center; font-size:40px; color:black;'>DMCH-IITD Machine Learning Tool for Estimating the Diet Percentage Similarity with Respect to Diets Consumed by Inflammatory Bowel Disease Patients Prior to Diagnosis</h1>",
        unsafe_allow_html=True
    )


with col_logo_right:
    st.markdown('<img src="https://raw.githubusercontent.com/gmaheshkumar15/ibd_similarity_score_-predictor3/main/dmch.jpeg" class="logo-right">', unsafe_allow_html=True)
    st.markdown('<div class="institute-name">Dayanand Medical College and Hospital Ludhiana</div>', unsafe_allow_html=True)

# -----------------------------
# Intro Paragraph
# -----------------------------
st.markdown("<hr style='border: 1px solid black;'>", unsafe_allow_html=True)
st.markdown("""
<p style='text-align:left; font-size:20px; color:black; line-height:1.5;'>
This tool is developed by DMCH Ludhiana and IIT Delhi. It uses machine learning (ML) models to estimate the similarity of a diet with those consumed by patients prior to an Inflammatory Bowel Disease (IBD) diagnosis. 
The ML model was trained based on data from a dietary survey conducted by DMCH Ludhiana among IBD patients and Controls without IBD. 
IBD patients were asked to report their dietary habits prior to diagnosis, and Controls were asked to report current food habits.
</p>
""", unsafe_allow_html=True)
st.markdown("<hr style='border: 1px solid black;'>", unsafe_allow_html=True)

# -----------------------------
# NEW SECTION — Download Excel Template
# -----------------------------
st.markdown("Download the Excel template, select the consumption level for each food item (higher values indicate higher consumption), and upload the file below to get predictions.")
try:
    with open("DMCH-IITD.xlsx", "rb") as f:
        st.download_button(
            label="Download Excel Template",
            data=f,
            file_name="DMCH-IITD.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
except FileNotFoundError:
    st.warning("⚠️ Excel template file not found. Please ensure 'DMCH-IITD.xlsx' is in the same folder as app.py")

st.markdown("<hr style='border: 1px solid black;'>", unsafe_allow_html=True)

# -----------------------------
# File Upload Section
# -----------------------------
uploaded_file = st.file_uploader("Upload Excel file with 77 features", type=["xlsx"])

if uploaded_file:
    try:
        df_raw = pd.read_excel(uploaded_file, engine="openpyxl")
        st.subheader("Input Features")
        st.dataframe(df_raw.head())

        # Step 1: Merge raw features → 22 merged features
        df_merged = merge_features(df_raw)

        # Step 2: Scale merged features
        df_scaled = scaler.transform(df_merged)

        # Step 3: Predictions
        log_prob = log_model.predict_proba(df_scaled)[:, 1]

        df_predictions = df_merged.copy()
        df_predictions["Logistic_Prob"] = log_prob

        # Step 4: Layout — Left (features) | Right (predictions)
        col_left, col_right = st.columns([2, 1])

        with col_left:
            st.subheader("Merged 22 Features")

            feature_names = df_merged.columns.tolist()
            first_half = feature_names[:11]
            second_half = feature_names[11:]

            c1, c2 = st.columns(2)

            df_display1 = pd.DataFrame({
                "No.": range(1, len(first_half) + 1),
                "Feature": first_half,
                "Value": [df_merged.iloc[0][f] for f in first_half]
            })
            df_display2 = pd.DataFrame({
                "No.": range(len(first_half) + 1, len(first_half) + len(second_half) + 1),
                "Feature": second_half,
                "Value": [df_merged.iloc[0][f] for f in second_half]
            })

            df_display1["Value"] = df_display1["Value"].round(0).astype(int)
            df_display2["Value"] = df_display2["Value"].round(0).astype(int)


            df_display1 = df_display1.reset_index(drop=True)
            df_display2 = df_display2.reset_index(drop=True)

            styler1 = df_display1.style.set_table_styles([
                {"selector": "th", "props": [("font-weight", "bold"), ("text-align", "center")]},
                {"selector": "td", "props": [("text-align", "center")]}
            ])
            styler2 = df_display2.style.set_table_styles([
                {"selector": "th", "props": [("font-weight", "bold"), ("text-align", "center")]},
                {"selector": "td", "props": [("text-align", "center")]}
            ])

            with c1:
                st.dataframe(styler1, use_container_width=True, hide_index=True)
            with c2:
                st.dataframe(styler2, use_container_width=True, hide_index=True)

        with col_right:
            st.markdown("<h3 style='text-align:center;'>Similarity Score</h3>", unsafe_allow_html=True)
            st.markdown(f"<p style='font-size:50px; font-weight:bold; color:#8B0000; text-align:center;'>{log_prob[0] * 100:.0f}</p>",unsafe_allow_html=True)

            
            # ---------- Step 5: Formatted Excel (Vertical Layout with Title & Borders) ----------
            output = BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "Prediction Results"

            # ---- Title Row ----
            ws.merge_cells("A1:B1")
            title_cell = ws["A1"]
            title_cell.value = "IBD Prediction Results"
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal="center", vertical="center")

            # ---- Header Row ----
            ws.append(["Feature", "Value"])
            ws["A2"].font = Font(bold=True)
            ws["B2"].font = Font(bold=True)
            ws["A2"].alignment = ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

            # ---- Add Feature Rows ----
            for f_name in df_merged.columns:
                ws.append([f_name, df_merged.iloc[0][f_name]])

            # ---- Blank Row ----
            ws.append([])

            # ---- Add Model Results ----
            ws.append(["Model", "Similarity (%)"])
            ws["A{}".format(ws.max_row)].font = Font(bold=True)
            ws["B{}".format(ws.max_row)].font = Font(bold=True)
            ws["A{}".format(ws.max_row)].alignment = ws["B{}".format(ws.max_row)].alignment = Alignment(horizontal="center", vertical="center")

            ws.append(["Logistic Regression", round(log_prob[0] * 100, 0)])
            
            # ---- Borders (inside + outside) ----
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # ---- Auto Column Width ----
            for col in ["A", "B"]:
                max_len = max(len(str(cell.value)) if cell.value else 0 for cell in ws[col])
                ws.column_dimensions[col].width = max_len + 4

            wb.save(output)
            output.seek(0)

            st.download_button(
                label="Download Results as Excel",
                data=output,
                file_name="IBD_Prediction_Results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    except Exception as e:
        st.error(f"Error processing file: {e}")
